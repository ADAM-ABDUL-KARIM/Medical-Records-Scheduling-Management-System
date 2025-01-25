from django.shortcuts import render
from django.contrib.auth.models import User
from rest_framework import generics
from .serializers import *
from rest_framework.permissions import AllowAny, IsAuthenticated
from .models import *
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import JsonResponse

class HealthcareProfessionalRetrieve(generics.ListCreateAPIView):
    serializer_class = HealthCareProfessionalSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return HealthCareProfessional.objects.all()

    def perform_create(self, serializer):
        return serializer.save()

class PatientRetrieve(generics.ListCreateAPIView):
    serializer_class = PatientSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        user = self.request.user
        if user.groups.filter(name='Patient').exists():
            return Patient.objects.filter(user=user)
        return Patient.objects.all()
    
    def perform_create(self, serializer):
        serializer.save()

class PatientUpdate(generics.UpdateAPIView):
    serializer_class = PatientUpdateSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return Patient.objects.all()

class PatientDelete(generics.DestroyAPIView):
    serializer_class = PatientSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return Patient.objects.all()

class AppointmentCreate(generics.ListCreateAPIView):
    serializer_class = AppointmentSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        user = self.request.user
        if user.groups.filter(name='Patient').exists():
            return Appointment.objects.filter(patient__user=user)
        return Appointment.objects.all()

    def perform_create(self, serializer):
        if serializer.is_valid():
            serializer.save()
        else:
            print(serializer.errors)
        return super()

class AppointmentDelete(generics.DestroyAPIView):
    serializer_class = AppointmentSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return Appointment.objects.all()

class NoteListCreate(generics.ListCreateAPIView):
    serializer_class = NoteSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.groups.filter(name='Patient').exists():
            return Note.objects.filter(patient__user=user)
        return Note.objects.all()

class NoteDelete(generics.DestroyAPIView):
    serializer_class = NoteSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return Note.objects.all()

class CreateUserView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        username = request.data.get('username')
        is_superuser = request.data.get('is_superuser', False)
        if User.objects.filter(username=username).exists():
            return Response({"detail": "Username already exists."}, status=status.HTTP_409_CONFLICT)
        
        user = User(
            username=username,
            is_superuser=is_superuser,
            is_staff=is_superuser
        )
        user.set_password(request.data.get('password'))
        user.save()
        
        return Response({"detail": "User created successfully."}, status=status.HTTP_201_CREATED)

class ProfileView(generics.RetrieveAPIView):
    queryset = Profile.objects.all()
    serializer_class = ProfileSerializer
    permission_classes = [AllowAny]

    def get_object(self):
        return self.request.user.profile

class UsernameView(generics.RetrieveAPIView):
    permission_classes = [AllowAny]

    def get(self, request):
        if request.user.is_authenticated:
            is_patient = request.user.groups.filter(name='Patient').exists()
            return Response({"username": request.user.username, "is_patient": is_patient})
        return Response({"username": "Admin", "is_patient": False})

class AvailabilityView(generics.ListCreateAPIView):
    serializer_class = AvailabilitySerializer
    permission_classes = [AllowAny]

    def perform_create(self, serializer):
        if serializer.is_valid():
            serializer.save()
        else:
            print(serializer.errors)
        return super()

    def get_queryset(self):
        return Availability.objects.all()

class AvailabilityDelete(generics.DestroyAPIView):
    serializer_class = AvailabilitySerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return Availability.objects.all()

def export_patients_excel(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Patients'
    columns = [
        'First Name', 'Last Name', 'Date Of Birth', 'Nationality', 'Address',
        'Marital Status', 'Phone Number', 'Gender', 'Height', 'Educational Level',
        'Employment Status', 'Dominant Hand', 'Start Date', 'Activity Level', 'Is Recovered',
        'Diagnosis', 'Medication', 'Notes'
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    diagnosis = ", ".join([d.diagnosis for d in patient.diagnosis.all()])
    medication = ", ".join([m.medication_name for m in patient.medication.all()])
    notes = ", ".join([n.note_content for n in patient.notes.all()])

    patient_data = [
        patient.first_name, patient.last_name, patient.dob, patient.nationality, patient.address,
        patient.marital_status, patient.phone_number, patient.gender, patient.height, patient.educational_level,
        patient.employment_status, patient.dominant_hand, patient.start_date, patient.activity_level, patient.is_recovered,
        diagnosis, medication, notes
    ]
    row_num += 1
    for col_num, cell_value in enumerate(patient_data, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value

    filename = f'patient_{patient.first_name}_{patient.last_name}_{pk}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response

def export_patient_pdf(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    response = HttpResponse(content_type='application/pdf')
    filename = f'patient_{patient.first_name}_{patient.last_name}_{pk}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    p = canvas.Canvas(response)
    width, height = p._pagesize  # Get page dimensions

    def draw_patient_info(p, patient, y, height):
        fields = [
            f"Patient Report for {patient.first_name} {patient.last_name}",
            f"First Name: {patient.first_name}",
            f"Last Name: {patient.last_name}",
            f"Date of Birth: {patient.dob}",
            f"Nationality: {patient.nationality}",
            f"Address: {patient.address}",
            f"Marital Status: {patient.marital_status}",
            f"Phone Number: {patient.phone_number}",
            f"Gender: {patient.gender}",
            f"Height: {patient.height} cm",
            f"Educational Level: {patient.educational_level}",
            f"Employment Status: {patient.employment_status}",
            f"Dominant Hand: {patient.dominant_hand}",
            f"Start Date: {patient.start_date}",
            f"Activity Level: {patient.activity_level}",
            f"Recovered: {'Yes' if patient.is_recovered else 'No'}",
            f"Diagnosis: {', '.join([d.diagnosis for d in patient.diagnosis.all()])}",
            f"Medication: {', '.join([m.medication_name for m in patient.medication.all()])}",
            f"Notes: {', '.join([n.note_content for n in patient.notes.all()])}"
        ]

        # Loop through fields and draw them on the canvas
        for field in fields:
            if y < 50:  # If space is less than the bottom margin
                p.showPage()  # Create a new page
                y = height - 100  # Reset the y-coordinate for the new page
            p.drawString(100, y, field)  # Draw the text
            y -= 50  # Move down by 50 units for the next line

        return y

    # Initial vertical position
    y = height - 100
    y = draw_patient_info(p, patient, y, height)

    # Save the PDF and return the response
    p.showPage()
    p.save()
    return response





def analytics_view(request):
    total_patients = Patient.objects.count()
    recovered_patients = Patient.objects.filter(is_recovered=True).count()
    not_recovered_patients = total_patients - recovered_patients

    data = {
        "total_patients": total_patients,
        "recovered_patients": recovered_patients,
        "not_recovered_patients": not_recovered_patients,
    }
    return JsonResponse(data)